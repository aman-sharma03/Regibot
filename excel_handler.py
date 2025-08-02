#!/usr/bin/env python3
"""
Excel File Handler - Comprehensive example for reading and writing Excel files

This script demonstrates various ways to work with Excel files:
1. Creating sample Excel files
2. Reading Excel files with different methods
3. Processing and analyzing Excel data
4. Writing data back to Excel files

Dependencies: pandas, openpyxl, xlrd
Install with: pip install pandas openpyxl xlrd
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime, timedelta


class ExcelHandler:
    def __init__(self):
        self.data = None
    
    def create_sample_excel(self, filename="sample_data.xlsx"):
        """Create a sample Excel file with multiple sheets for demonstration"""
        
        # Create sample data
        employees_data = {
            'Employee_ID': [1001, 1002, 1003, 1004, 1005, 1006, 1007, 1008],
            'Name': ['John Smith', 'Jane Doe', 'Mike Johnson', 'Sarah Wilson', 
                    'David Brown', 'Lisa Garcia', 'Tom Anderson', 'Amy Taylor'],
            'Department': ['IT', 'HR', 'Finance', 'IT', 'Marketing', 'Finance', 'HR', 'IT'],
            'Salary': [75000, 65000, 70000, 80000, 60000, 72000, 58000, 78000],
            'Hire_Date': ['2020-01-15', '2019-03-22', '2021-07-10', '2018-11-05',
                         '2022-02-18', '2020-09-30', '2019-12-12', '2021-04-25'],
            'Performance_Score': [4.2, 3.8, 4.5, 4.0, 3.9, 4.3, 3.7, 4.1]
        }
        
        sales_data = {
            'Date': pd.date_range('2024-01-01', periods=12, freq='M'),
            'Product_A': [15000, 18000, 22000, 19000, 25000, 28000, 
                         30000, 27000, 32000, 35000, 38000, 40000],
            'Product_B': [12000, 14000, 16000, 15000, 18000, 20000,
                         22000, 21000, 24000, 26000, 28000, 30000],
            'Product_C': [8000, 9500, 11000, 10500, 12000, 13500,
                         15000, 14000, 16000, 17500, 19000, 21000]
        }
        
        # Create DataFrames
        employees_df = pd.DataFrame(employees_data)
        sales_df = pd.DataFrame(sales_data)
        
        # Write to Excel with multiple sheets
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            employees_df.to_excel(writer, sheet_name='Employees', index=False)
            sales_df.to_excel(writer, sheet_name='Sales', index=False)
            
            # Add a summary sheet
            summary_data = {
                'Metric': ['Total Employees', 'Average Salary', 'Total Departments',
                          'Latest Sales Month', 'Best Performing Product'],
                'Value': [len(employees_df), 
                         f"${employees_df['Salary'].mean():,.0f}",
                         employees_df['Department'].nunique(),
                         sales_df['Date'].max().strftime('%Y-%m'),
                         'Product_A']
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        print(f"✅ Sample Excel file '{filename}' created successfully!")
        return filename
    
    def read_excel_basic(self, filename):
        """Basic method to read Excel file using pandas"""
        try:
            # Read the first sheet (default behavior)
            df = pd.read_excel(filename)
            print(f"📊 Basic read - Shape: {df.shape}")
            print(f"Columns: {list(df.columns)}")
            print("\nFirst 3 rows:")
            print(df.head(3))
            return df
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
            return None
    
    def read_excel_advanced(self, filename):
        """Advanced method to read Excel file with multiple sheets"""
        try:
            # Read all sheets
            excel_file = pd.ExcelFile(filename)
            print(f"📋 Available sheets: {excel_file.sheet_names}")
            
            all_data = {}
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(filename, sheet_name=sheet_name)
                all_data[sheet_name] = df
                print(f"\n📊 Sheet '{sheet_name}' - Shape: {df.shape}")
                print(f"Columns: {list(df.columns)}")
            
            return all_data
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
            return None
    
    def read_excel_with_options(self, filename, sheet_name=0, **kwargs):
        """Read Excel with various options"""
        try:
            # Common options for reading Excel files
            options = {
                'sheet_name': sheet_name,
                'header': kwargs.get('header', 0),  # Which row to use as column names
                'skiprows': kwargs.get('skiprows', None),  # Rows to skip
                'usecols': kwargs.get('usecols', None),  # Columns to read
                'nrows': kwargs.get('nrows', None),  # Number of rows to read
                'dtype': kwargs.get('dtype', None),  # Data types
                'parse_dates': kwargs.get('parse_dates', None),  # Columns to parse as dates
            }
            
            # Remove None values
            options = {k: v for k, v in options.items() if v is not None}
            
            df = pd.read_excel(filename, **options)
            print(f"📊 Advanced read with options - Shape: {df.shape}")
            return df
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
            return None
    
    def analyze_data(self, df, sheet_name="Data"):
        """Perform basic analysis on the data"""
        if df is None or df.empty:
            print("❌ No data to analyze")
            return
        
        print(f"\n🔍 Analysis for {sheet_name}:")
        print("=" * 50)
        
        # Basic info
        print(f"Shape: {df.shape}")
        print(f"Memory usage: {df.memory_usage(deep=True).sum() / 1024:.2f} KB")
        
        # Data types
        print(f"\nData types:")
        for col, dtype in df.dtypes.items():
            print(f"  {col}: {dtype}")
        
        # Missing values
        missing = df.isnull().sum()
        if missing.any():
            print(f"\nMissing values:")
            for col, count in missing[missing > 0].items():
                print(f"  {col}: {count}")
        else:
            print("\n✅ No missing values")
        
        # Numeric columns summary
        numeric_cols = df.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            print(f"\nNumeric columns summary:")
            print(df[numeric_cols].describe())
    
    def write_excel_with_formatting(self, data, filename="formatted_output.xlsx"):
        """Write data to Excel with formatting using openpyxl"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Formatted Data"
        
        # If data is a DataFrame, convert to list of lists
        if isinstance(data, pd.DataFrame):
            # Add headers
            headers = list(data.columns)
            ws.append(headers)
            
            # Add data
            for _, row in data.iterrows():
                ws.append(list(row))
        
        # Apply formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        print(f"✅ Formatted Excel file '{filename}' created successfully!")


def main():
    """Main function to demonstrate Excel file operations"""
    print("🚀 Excel File Handler Demo")
    print("=" * 50)
    
    handler = ExcelHandler()
    
    # 1. Create a sample Excel file
    print("\n1. Creating sample Excel file...")
    sample_file = handler.create_sample_excel("sample_data.xlsx")
    
    # 2. Basic reading
    print("\n2. Basic Excel reading...")
    df_basic = handler.read_excel_basic(sample_file)
    
    # 3. Advanced reading (all sheets)
    print("\n3. Advanced Excel reading (all sheets)...")
    all_sheets = handler.read_excel_advanced(sample_file)
    
    # 4. Reading with specific options
    print("\n4. Reading with options (first 5 rows of Employees sheet)...")
    df_options = handler.read_excel_with_options(
        sample_file, 
        sheet_name='Employees',
        nrows=5,
        usecols=['Name', 'Department', 'Salary']
    )
    print(df_options)
    
    # 5. Data analysis
    if all_sheets:
        for sheet_name, df in all_sheets.items():
            handler.analyze_data(df, sheet_name)
    
    # 6. Write formatted Excel
    print("\n6. Creating formatted Excel file...")
    if df_basic is not None:
        handler.write_excel_with_formatting(df_basic, "formatted_output.xlsx")
    
    print("\n✅ Demo completed! Check the created Excel files.")
    
    # List created files
    excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    print(f"\n📁 Excel files in current directory: {excel_files}")


if __name__ == "__main__":
    main()